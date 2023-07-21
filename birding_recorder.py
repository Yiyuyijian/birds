import aves
import openpyxl
import tkinter as tk
from tkinter import ttk
from openpyxl.styles import numbers,Font,Side,Alignment,Border,PatternFill,Color
import tkinter.messagebox
import calendar
from datetime import datetime

class Writer(object):
    def __init__(self) -> None:
        super().__init__()
    
        self.Myformat = self.Wt_excel.add_format({
                'font_size':  15,       # 字体大小
                'bold':       True,     # 是否粗体
                'bg_color':   '#b2b2b2',# 表格背景颜色
                'align':      'center', # 水平居中对齐
                'valign':     'vcenter',# 垂直居中对齐
                'text_wrap':  True,     # 设置自动换行
                'border':     1,        # 边框宽度
                'top':        1,        # 上边框
                'left':       1,        # 左边框
                'right':      1,        # 右边框
                'bottom':     1         # 底边框
                })
#基础类，提供主界面和容器
class Base_UI(object):
    def __init__(self):
        super().__init__()
        self.Win=tk.Tk()
        self.Win.title('观鸟记录小程序')
        self.Win.geometry('800x600')
        self.Win.resizable(0,0)#框体大小可调性，分别表示x,y方向的可变性

        self.excel=openpyxl.load_workbook('E:/animals/我的记录.xlsx')
        self.sheets=self.excel.worksheets

        tk.Label(self.Win,text='选择日期：',font=('隶书',20)).place(x=50,y=20)

        self.Year_combobox=ttk.Combobox(self.Win,width=10,values=list(range(2020,2070)))
        self.Year_combobox.place(x=210,y=25)
        tk.Label(self.Win,text='年',font=('隶书',15)).place(x=310,y=25)

        self.Month_combobox=ttk.Combobox(self.Win,width=10)
        self.Month_combobox['values']=list(range(1,13))
        self.Month_combobox.place(x=350,y=25)
        self.Month_combobox.bind("<<ComboboxSelected>>",self.__set_days)
        tk.Label(self.Win,text='月',font=('隶书',15)).place(x=450,y=25)

        self.Day_combobox=ttk.Combobox(self.Win,width=10)
        self.Day_combobox.place(x=490,y=25)
        tk.Label(self.Win,text='日',font=('隶书',15)).place(x=590,y=25)

        self.__aves=aves.Aves_China(5.0)

        tk.Label(self.Win,text="选择目：",font=('隶书',20)).place(x=50,y=100)
        self.Order_combobox=ttk.Combobox(self.Win,width=15,values=self.__aves.orders())
        self.Order_combobox.place(x=50,y=140)
        self.Order_combobox.bind("<<ComboboxSelected>>",self.__set_families)


        tk.Label(self.Win,text="选择科：",font=('隶书',20)).place(x=50,y=200)
        self.Family_combobox=ttk.Combobox(self.Win,width=15)
        self.Family_combobox.place(x=50,y=240)
        self.Family_combobox.bind("<<ComboboxSelected>>",self.__set_genuses)
        

        tk.Label(self.Win,text="选择属：",font=('隶书',20)).place(x=50,y=300)
        self.Genus_combobox=ttk.Combobox(self.Win,width=15)
        self.Genus_combobox.place(x=50,y=340)
        self.Genus_combobox.bind("<<ComboboxSelected>>",self.__set_species)

        tk.Label(self.Win,text="选择种：",font=('隶书',20)).place(x=50,y=400)
        self.Species_combobox=ttk.Combobox(self.Win,width=15)
        self.Species_combobox.place(x=50,y=440)

        tk.Label(self.Win,text='输入地点：',font=('隶书',20)).place(x=300,y=100)
        tk.Label(self.Win,text='省级行政单位：',font=('隶书',15)).place(x=300,y=150)
        self.Province_entry=tk.Entry(self.Win,width=20)
        self.Province_entry.place(x=300,y=180)

        tk.Label(self.Win,text='地级行政单位：',font=('隶书',15)).place(x=300,y=240)
        self.City_entry=tk.Entry(self.Win,width=20)
        self.City_entry.place(x=300,y=270)

        tk.Label(self.Win,text='县级行政单位：',font=('隶书',15)).place(x=300,y=330)
        self.County_entry=tk.Entry(self.Win,width=20)
        self.County_entry.place(x=300,y=360)

        tk.Label(self.Win,text='具体地点：',font=('隶书',15)).place(x=300,y=420)
        self.Local_entry=tk.Entry(self.Win,width=20)
        self.Local_entry.place(x=300,y=450)

        tk.Label(self.Win,text='备  注：',font=('隶书',20)).place(x=500,y=100)
        self.Note=tk.Text(self.Win,width=35,bd=5)
        self.Note.place(x=500,y=160)

        once=tk.Button(self.Win,text='录完一种点击',bg='red',width=20,height=1,command=self.__save_data)
        once.place(x=50,y=500)
        over=tk.Button(self.Win,text='全部录完点击',bg='red',width=20,height=1,command=self.__finish)
        over.place(x=300,y=500)
        
    def __set_days(self,*args):
        try:
            year = int(self.Year_combobox.get())
            month = int(self.Month_combobox.get())
            days = calendar.monthrange(year,month)[1]
            self.Day_combobox.delete(0,tk.END)
            self.Day_combobox.config(values=list(range(1,days+1)))
        except:
            self.Month_combobox.delete(0,tk.END)
            tkinter.messagebox.showerror(title="未选择年",message="请先选择年份")
            
    
    def __set_families(self,*args):
        self.Family_combobox['values']=self.__aves.families_in_order(self.Order_combobox.get())
        self.Genus_combobox['values']=[]
        self.Species_combobox['values']=[]
        self.Family_combobox.delete(0,tk.END)
        self.Genus_combobox.delete(0,tk.END)
        self.Species_combobox.delete(0,tk.END)

    def __set_genuses(self,*args):
        self.Genus_combobox['values']=self.__aves.genuses_in_family(self.Family_combobox.get())
        self.Species_combobox['values']=[]
        self.Genus_combobox.delete(0,tk.END)
        self.Species_combobox.delete(0,tk.END)

    def __set_species(self,*args):
        self.Species_combobox['values']=self.__aves.species_in_genus(self.Genus_combobox.get())
        self.Species_combobox.delete(0,tk.END)

    def __save_data(self):
        #组装数据
        datas = []
        datas.append(self.sheets[0].max_row)
        datas.append(self.Species_combobox.get())
        year = int(self.Year_combobox.get())
        month = int(self.Month_combobox.get())
        day = int(self.Day_combobox.get() )       
        datas.append(datetime(year,month,day))        
        datas.append(self.Province_entry.get())
        datas.append(self.City_entry.get())
        datas.append(self.County_entry.get())
        datas.append(self.Local_entry.get())
        datas.append(self.Order_combobox.get())
        datas.append(self.Family_combobox.get())
        datas.append(self.Genus_combobox.get())
        datas.append(self.Note.get(1.0,tk.END))

        #续写
        self.sheets[0].append(datas)
        #设置新添加的行的格式
        for cell in self.sheets[0][self.sheets[0].max_row]:
            cell.font=Font(name='宋体',size=12)
            cell.border=Border(left = Side(style="thin"),
                               right= Side(style="thin"),
                               top  = Side(style="thin"),
                               bottom=Side(style="thin"))
            cell.fill=PatternFill(bgColor=Color(index=3))
            cell.alignment=Alignment(horizontal='center',vertical='center')
        #设置日期显示格式
        self.sheets[0].cell(self.sheets[0].max_row,3).number_format=numbers.FORMAT_DATE_YYYYMMDD2
        #设置行高
        self.sheets[0].row_dimensions[self.sheets[0].max_row].height=20

        #清除上一次数据
        self.Species_combobox.delete(0,tk.END)
        self.Year_combobox.delete(0,tk.END)
        self.Month_combobox.delete(0,tk.END)
        self.Day_combobox.delete(0,tk.END)
        self.Province_entry.delete(0,tk.END)
        self.City_entry.delete(0,tk.END)
        self.County_entry.delete(0,tk.END)
        self.Local_entry.delete(0,tk.END)
        self.Order_combobox.delete(0,tk.END)
        self.Family_combobox.delete(0,tk.END)
        self.Genus_combobox.delete(0,tk.END)
        self.Note.delete(1.0,tk.END)
        tkinter.messagebox.showinfo('提示','一种鸟的数据已保存')
 
    def __finish(self):
        self.excel.save('E:/animals/我的记录.xlsx')
        tkinter.messagebox.showinfo('提示','所有录入的鸟种记录已保存')

if __name__ == "__main__":
    Base_UI().Win.mainloop()
